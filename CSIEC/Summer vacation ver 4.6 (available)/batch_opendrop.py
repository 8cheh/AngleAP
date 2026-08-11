#!/usr/bin/env python3
"""
接触角批量测量 — 使用 OpenDrop 算法
=====================================

基于 OpenDrop (https://github.com/opendrop-dev/opendrop) 的核心算法:
  - extract_contact_angle_features: 梯度边缘 + 自适应阈值 + Canny细化 + 连通域
  - contact_angle_fit: 基线坐标变换 → 左右分离 → 圆弧/直线拟合 → 接触角

参考文献:
  J. D. Berry et al., J. Colloid Interface Sci. 454 (2015) 226–237.
  E. Huang et al., OpenDrop, submitted to J. Open Source Software.

输出: opendrop测量结果.xlsx
"""

import os, sys, time, warnings, math
import numpy as np
import cv2
import scipy.optimize
from scipy.ndimage import uniform_filter1d
from enum import IntEnum, auto

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ===========================================================================
# 路径配置
# ===========================================================================
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGE_DIR = os.path.join(_SCRIPT_DIR, '图片', '图片')
OUTPUT_XLSX = os.path.join(_SCRIPT_DIR, 'opendrop测量结果.xlsx')

# ===========================================================================
# OpenDrop Geometry (geometry.py)
# ===========================================================================
PI = math.pi


class Vector2:
    """2D vector, compatible with OpenDrop's Vector2."""
    __slots__ = ('x', 'y')

    def __init__(self, x=0.0, y=0.0):
        if hasattr(x, '__iter__') and not isinstance(x, (str, bytes)):
            x, y = x
        self.x = float(x)
        self.y = float(y)

    def __getitem__(self, i):
        return self.x if i == 0 else self.y

    def __len__(self):
        return 2

    def __iter__(self):
        yield self.x
        yield self.y

    def __add__(self, o):
        o = Vector2(o) if not isinstance(o, Vector2) else o
        return Vector2(self.x + o.x, self.y + o.y)

    def __sub__(self, o):
        o = Vector2(o) if not isinstance(o, Vector2) else o
        return Vector2(self.x - o.x, self.y - o.y)

    def __mul__(self, s):
        if isinstance(s, (int, float)):
            return Vector2(self.x * s, self.y * s)
        s = Vector2(s) if not isinstance(s, Vector2) else s
        return Vector2(self.x * s.x, self.y * s.y)

    def __rmul__(self, s):
        return self.__mul__(s)

    def __truediv__(self, s):
        if isinstance(s, (int, float)):
            return Vector2(self.x / s, self.y / s)
        s = Vector2(s) if not isinstance(s, Vector2) else s
        return Vector2(self.x / s.x, self.y / s.y)

    def __neg__(self):
        return Vector2(-self.x, -self.y)

    def __abs__(self):
        return math.hypot(self.x, self.y)

    def __matmul__(self, o):
        """Dot product. Works with Vector2, (2,) array, or (2, N) array."""
        if isinstance(o, Vector2):
            return self.x * o.x + self.y * o.y
        # numpy array: (2,) or (2, N) -> broadcasting dot
        o = np.asarray(o)
        if o.ndim == 1:
            return self.x * o[0] + self.y * o[1]
        elif o.ndim == 2:
            return self.x * o[0] + self.y * o[1]
        raise ValueError(f"Expected (2,) or (2,N) array, got shape {o.shape}")

    def __repr__(self):
        return f'Vector2({self.x}, {self.y})'

    def __eq__(self, o):
        if not isinstance(o, Vector2):
            return False
        return self.x == o.x and self.y == o.y


class Rect2:
    """Axis-aligned rectangle, compatible with OpenDrop's Rect2."""
    __slots__ = ('_x0', '_y0', '_x1', '_y1')

    def __init__(self, x0=0, y0=0, x1=0, y1=0):
        self._x0 = int(x0)
        self._y0 = int(y0)
        self._x1 = int(x1)
        self._y1 = int(y1)

    @property
    def x0(self): return self._x0

    @property
    def y0(self): return self._y0

    @property
    def x1(self): return self._x1

    @property
    def y1(self): return self._y1

    @property
    def pt0(self): return Vector2(self.x0, self.y0)

    @property
    def pt1(self): return Vector2(self.x1, self.y1)

    @property
    def position(self): return Vector2(self.x0, self.y0)

    @property
    def w(self): return self.x1 - self.x0

    @property
    def h(self): return self.y1 - self.y0

    def __repr__(self):
        return f'Rect2(x0={self.x0}, y0={self.y0}, x1={self.x1}, y1={self.y1})'


class Line2:
    """Line defined by two points, compatible with OpenDrop's Line2."""
    __slots__ = ('_pt0', '_pt1')

    def __init__(self, pt0, pt1):
        self._pt0 = Vector2(pt0)
        self._pt1 = Vector2(pt1)
        if abs(self._pt1 - self._pt0) < 1e-12:
            raise ValueError("pt0 and pt1 cannot be equal")

    @property
    def pt0(self): return self._pt0

    @property
    def pt1(self): return self._pt1

    @property
    def unit(self):
        u = self.pt1 - self.pt0
        return u / abs(u)

    @property
    def perp(self):
        u = self.unit
        return Vector2(-u.y, u.x)

    @property
    def gradient(self):
        dx = self._pt1.x - self._pt0.x
        if dx == 0:
            return math.copysign(float('inf'), self._pt1.y - self._pt0.y)
        return (self._pt1.y - self._pt0.y) / dx

    @property
    def angle(self):
        u = self.unit
        return math.atan2(u.y, u.x)

    def eval(self, *, x=None, y=None):
        if x is not None:
            return Vector2(x, self._solve_for_y(x))
        elif y is not None:
            return Vector2(self._solve_for_x(y), y)
        raise TypeError

    def solve(self, *, x=None, y=None):
        if x is not None:
            return self._solve_for_y(x)
        elif y is not None:
            return self._solve_for_x(y)
        raise TypeError

    def _solve_for_y(self, x):
        return self._pt0.y + (x - self._pt0.x) * self.gradient

    def _solve_for_x(self, y):
        if self.gradient == 0:
            return math.copysign(float('inf'), y - self._pt0.y)
        return self._pt0.x + (y - self._pt0.y) / self.gradient

    def __add__(self, t):
        t = Vector2(t)
        return Line2(self.pt0 + t, self.pt1 + t)

    def __sub__(self, t):
        t = Vector2(t)
        return Line2(self.pt0 - t, self.pt1 - t)

    def __repr__(self):
        return f'Line2({tuple(self.pt0)}, {tuple(self.pt1)})'


def rotation_mat2d(theta):
    """2D rotation matrix."""
    c, s = math.cos(theta), math.sin(theta)
    return np.array([[c, -s], [s, c]])


# ===========================================================================
# OpenDrop Line Fitting (fit/line/)
# ===========================================================================
class LineParam(IntEnum):
    ANGLE = 0
    RHO = auto()


class LineModel:
    def __init__(self, data):
        self.data = data
        self._params = np.empty(2)
        self._residuals = np.empty(data.shape[1])
        self._jac = np.empty((data.shape[1], 2))

    def set_params(self, params):
        q, rho = params[LineParam.ANGLE], params[LineParam.RHO]
        Q = rotation_mat2d(q)
        r, z = Q.T @ self.data - [[0], [rho]]
        self._residuals[:] = z
        self._jac[:, LineParam.ANGLE] = -r
        self._jac[:, LineParam.RHO] = -1
        self._params[:] = params

    @property
    def params(self): return self._params.copy()

    @property
    def dof(self): return max(1, self.data.shape[1] - 2 + 1)

    @property
    def jac(self): return self._jac.copy()

    @property
    def residuals(self): return self._residuals.copy()


def line_fit(data, *, loss='linear', f_scale=1.0, verbose=False):
    """Robust line fitting (identical to OpenDrop's line_fit)."""
    if data.shape[1] < 2:
        return None
    model = LineModel(data)

    def fun(params):
        model.set_params(params)
        return model.residuals.copy()

    def jac(params):
        model.set_params(params)
        return model.jac.copy()

    # Initial guess: line through first and last points
    pt0, pt1 = data[:, 0], data[:, -1]
    if np.all(pt0 == pt1):
        init = np.array([0.0, 0.0])
    else:
        line = Line2(pt0, pt1)
        init = np.array([line.angle, line.perp @ line.pt0])

    model.set_params(init)
    try:
        res = scipy.optimize.least_squares(
            fun, model.params, jac,
            method='lm' if loss == 'linear' else 'trf',
            loss=loss, f_scale=f_scale, x_scale='jac',
            ftol=1e-8, xtol=1e-8, gtol=1e-8, max_nfev=50,
            verbose=2 if verbose else 0,
        )
    except ValueError:
        return None
    model.set_params(res.x)
    return {
        'angle': model.params[LineParam.ANGLE],
        'rho': model.params[LineParam.RHO],
        'objective': (model.residuals ** 2).sum() / model.dof,
        'residuals': model.residuals.copy(),
    }


# ===========================================================================
# OpenDrop Circle Fitting (fit/circle/)
# ===========================================================================
class CircleParam(IntEnum):
    CENTER_X = 0
    CENTER_Y = auto()
    RADIUS = auto()


class CircleModel:
    def __init__(self, data):
        self.data = data
        self._params = np.empty(3)
        self._residuals = np.empty(data.shape[1])
        self._jac = np.empty((data.shape[1], 3))

    def set_params(self, params):
        xc, yc, R = params[CircleParam.CENTER_X], params[CircleParam.CENTER_Y], params[CircleParam.RADIUS]
        tx = self.data[0] - xc
        ty = self.data[1] - yc
        r = np.sqrt(tx ** 2 + ty ** 2)
        self._residuals[:] = r - R
        self._jac[:, CircleParam.CENTER_X] = -tx / r
        self._jac[:, CircleParam.CENTER_Y] = -ty / r
        self._jac[:, CircleParam.RADIUS] = -1
        self._params[:] = params

    @property
    def params(self): return self._params.copy()

    @property
    def dof(self): return max(1, self.data.shape[1] - 3 + 1)

    @property
    def jac(self): return self._jac.copy()

    @property
    def residuals(self): return self._residuals.copy()


def circle_fit(data, *, loss='linear', f_scale=1.0, xc=None, yc=None, radius=None, verbose=False):
    """Robust circle fitting (identical to OpenDrop's circle_fit)."""
    if data.shape[1] == 0:
        return None
    model = CircleModel(data)

    def fun(params):
        model.set_params(params)
        return model.residuals.copy()

    def jac(params):
        model.set_params(params)
        return model.jac.copy()

    init = np.empty(3)
    if xc is None or yc is None:
        xc_g, yc_g = data.mean(axis=1)
    else:
        xc_g, yc_g = xc, yc
    if radius is None:
        tx, ty = data[0] - xc_g, data[1] - yc_g
        radius_g = np.median(np.sqrt(tx ** 2 + ty ** 2))
    else:
        radius_g = radius

    init[CircleParam.CENTER_X] = xc_g
    init[CircleParam.CENTER_Y] = yc_g
    init[CircleParam.RADIUS] = radius_g
    model.set_params(init)

    try:
        res = scipy.optimize.least_squares(
            fun, model.params, jac,
            method='lm' if loss == 'linear' else 'trf',
            loss=loss, f_scale=f_scale, x_scale='jac',
            ftol=1e-8, xtol=1e-8, gtol=1e-8, max_nfev=50,
            verbose=2 if verbose else 0,
        )
    except ValueError:
        return None
    model.set_params(res.x)
    return {
        'center': Vector2(model.params[CircleParam.CENTER_X], model.params[CircleParam.CENTER_Y]),
        'radius': model.params[CircleParam.RADIUS],
        'objective': (model.residuals ** 2).sum() / model.dof,
        'residuals': model.residuals.copy(),
    }


# ===========================================================================
# OpenDrop Feature Extraction (features/conan.py)
# ===========================================================================
def extract_contact_angle_features(image, baseline, inverted, *, roi=None, thresh=0.3):
    """
    Extract drop edge points using OpenDrop's algorithm.

    Parameters
    ----------
    image : np.ndarray (H, W, 3) or (H, W)
        Input image (BGR or grayscale).
    baseline : Line2
        The baseline (surface line).
    inverted : bool
        False for sessile drop (drop above baseline in image coords).
    roi : Rect2 or None
        Region of interest.
    thresh : float
        Gradient threshold fraction (default 0.3 in OpenDrop).

    Returns
    -------
    drop_points : np.ndarray (2, N)
        Edge points of the drop profile.
    """
    if roi is None:
        roi = Rect2(0, 0, image.shape[1] - 1, image.shape[0] - 1)

    # Clip roi to image extents
    roi = Rect2(
        max(0, roi.x0), max(0, roi.y0),
        min(image.shape[1] - 1, roi.x1), min(image.shape[0] - 1, roi.y1),
    )

    subimage = image[roi.y0:roi.y1 + 1, roi.x0:roi.x1 + 1]
    baseline_shifted = baseline - roi.position

    if baseline_shifted.pt1.x < baseline_shifted.pt0.x:
        baseline_shifted = Line2(baseline_shifted.pt1, baseline_shifted.pt0)

    if inverted:
        up_vec = baseline_shifted.perp
        right_vec = baseline_shifted.unit
    else:
        # Drop is above baseline in image coords (negative y direction)
        up_vec = -baseline_shifted.perp
        right_vec = baseline_shifted.unit

    origin = np.array(baseline_shifted.pt0)

    if len(subimage.shape) > 2:
        gray = cv2.cvtColor(subimage, cv2.COLOR_RGB2GRAY)
    else:
        gray = subimage

    # OpenDrop edge detection pipeline
    blur = cv2.GaussianBlur(gray, ksize=(5, 5), sigmaX=0)
    dx = cv2.Scharr(blur, cv2.CV_16S, dx=1, dy=0)
    dy = cv2.Scharr(blur, cv2.CV_16S, dx=0, dy=1)

    # Gradient magnitude squared for sharper edges
    mask = (dx.astype(float) ** 2 + dy.astype(float) ** 2)
    mask = np.sqrt(mask)
    mask = (mask / mask.max() * 255).astype(np.uint8)

    # Ignore weak gradients
    mask[mask < thresh * mask.max()] = 0

    # Adaptive threshold
    cv2.adaptiveThreshold(
        mask, maxValue=1,
        adaptiveMethod=cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        thresholdType=cv2.THRESH_BINARY,
        blockSize=5, C=0, dst=mask,
    )

    # Remove edges below and within 2 pixels of the baseline
    mask_ij = np.array(mask.nonzero())
    if mask_ij.shape[1] > 0:
        y_up = up_vec @ (mask_ij[::-1] - origin.reshape(2, 1))
        ix = y_up.argsort()
        mask_ij = mask_ij[:, ix]
        y_up = y_up[ix]
        stop = np.searchsorted(y_up, 2.0, side='right')
        mask[tuple(mask_ij[:, :stop])] = 0

    # Connected components: keep the 2 largest
    n_labels, cc_labels, cc_stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=4)
    if n_labels > 1:
        sizes = cc_stats[:, cv2.CC_STAT_WIDTH] * cc_stats[:, cv2.CC_STAT_HEIGHT]
        ix = np.argsort(sizes)[::-1]
        ix = ix[:3]
        ix = ix[ix != 0]  # exclude background (label 0)

        if len(ix) >= 2:
            ext0 = Rect2(
                cc_stats[ix[0], cv2.CC_STAT_LEFT], cc_stats[ix[0], cv2.CC_STAT_TOP],
                cc_stats[ix[0], cv2.CC_STAT_LEFT] + cc_stats[ix[0], cv2.CC_STAT_WIDTH],
                cc_stats[ix[0], cv2.CC_STAT_TOP] + cc_stats[ix[0], cv2.CC_STAT_HEIGHT],
            )
            ext1 = Rect2(
                cc_stats[ix[1], cv2.CC_STAT_LEFT], cc_stats[ix[1], cv2.CC_STAT_TOP],
                cc_stats[ix[1], cv2.CC_STAT_LEFT] + cc_stats[ix[1], cv2.CC_STAT_WIDTH],
                cc_stats[ix[1], cv2.CC_STAT_TOP] + cc_stats[ix[1], cv2.CC_STAT_HEIGHT],
            )
            # If one is inside the other or much larger, use only the outer
            if (ext1.x0 > ext0.x0 and ext1.y0 > ext0.y0 and
                    ext1.x1 < ext0.x1 and ext1.y1 < ext0.y1) or \
                    ext0.w * ext0.h > 10 * ext1.w * ext1.h:
                mask &= cc_labels == ix[0]
            else:
                mask &= (cc_labels == ix[0]) | (cc_labels == ix[1])
        elif len(ix) == 1:
            mask &= cc_labels == ix[0]

    # Canny edge thinning
    grad_max = (abs(dx) + abs(dy)).max()
    if grad_max > 0:
        drop_edges = cv2.Canny(
            mask.astype(np.uint8) * dx.astype(np.int16),
            mask.astype(np.uint8) * dy.astype(np.int16),
            grad_max * thresh / 2, grad_max * thresh,
        )
    else:
        drop_edges = np.zeros_like(mask, dtype=np.uint8)

    drop_points_raw = np.array(drop_edges.nonzero()[::-1])

    if drop_points_raw.shape[1] == 0:
        return np.empty((2, 0), dtype=int)

    # Level-set based left/right edge extraction
    x_c, y_c = right_vec @ (drop_points_raw - origin.reshape(2, 1)), \
        up_vec @ (drop_points_raw - origin.reshape(2, 1))

    ix_sort = y_c.argsort()
    x_c, y_c = x_c[ix_sort], y_c[ix_sort]
    drop_points_raw = drop_points_raw[:, ix_sort]

    # Divide into ~2 pixel high level sets
    bins = max(1, int(y_c.max() / 2))
    if bins > 1:
        levels = np.histogram_bin_edges(y_c, bins=bins)
        levels_ix = [0] + list(np.searchsorted(y_c, levels[1:], side='right'))
    else:
        levels_ix = [0, len(y_c)]

    keep_mask = np.zeros(len(y_c), dtype=bool)
    for start, stop in zip(levels_ix, levels_ix[1:]):
        if stop <= start:
            continue
        level_set = x_c[start:stop]
        ltr_ix = level_set.argsort()
        # Split into contiguous groups (gap > 2*sqrt(2) ≈ 2.828)
        splits = (np.diff(level_set[ltr_ix]) > 2.828).nonzero()[0] + 1
        groups = np.split(ltr_ix, splits)
        if len(groups) >= 2:
            keep_mask[start + groups[0]] = True
            keep_mask[start + groups[-1]] = True
        elif len(groups) == 1:
            keep_mask[start + groups[0]] = True

    drop_points = drop_points_raw[:, keep_mask]

    # Shift back to full image coordinates
    drop_points = drop_points + np.reshape(roi.position, (2, 1))

    return drop_points


# ===========================================================================
# OpenDrop Contact Angle Fit (fit/conan.py)
# ===========================================================================
def contact_angle_fit_opendrop(data, baseline):
    """
    OpenDrop contact angle fitting algorithm.

    Parameters
    ----------
    data : np.ndarray (2, N)
        Drop edge points in image coordinates.
    baseline : Line2
        The baseline line.

    Returns
    -------
    dict with left_angle, right_angle, left_curvature, right_curvature, etc.
    """
    xy = data.astype(float)
    Q = np.array([baseline.unit, baseline.perp])

    # Transform to baseline coordinates (r, z)
    rz = Q @ (xy - np.reshape(baseline.pt0, (2, 1)))

    # Auto-detect which side the drop is on
    if abs(rz[1].min()) > abs(rz[1].max()):
        Q[1] *= -1
        rz[1] *= -1

    # Sort by height from baseline
    z_ix = rz[1].argsort()
    z_ix_inv = z_ix.argsort()
    rz = rz[:, z_ix]
    xy = xy[:, z_ix]

    rc = rz[0].mean()
    left_mask = rz[0] < rc
    right_mask = ~left_mask

    left_rz = rz[:, left_mask]
    right_rz = rz[:, right_mask]

    # Initial contact point guesses
    left_contact_rz = left_rz[:, 0] if left_rz.shape[1] > 0 else None
    right_contact_rz = right_rz[:, 0] if right_rz.shape[1] > 0 else None

    result = {
        'left_angle': None, 'right_angle': None,
        'left_curvature': None, 'right_curvature': None,
        'left_contact': None, 'right_contact': None,
        'left_arc_center': None, 'right_arc_center': None,
    }

    if left_contact_rz is not None and right_contact_rz is not None:
        base_width = right_contact_rz[0] - left_contact_rz[0]

        # Refine contact region: points within 25% of base_width from contact
        left_dists = np.linalg.norm(left_rz - np.reshape(left_contact_rz, (2, 1)), axis=0)
        right_dists = np.linalg.norm(right_rz - np.reshape(right_contact_rz, (2, 1)), axis=0)

        left_near = left_dists < 0.25 * base_width
        right_near = right_dists < 0.25 * base_width

        left_fit = _arc_fit(left_rz[:, left_near] if left_near.sum() > 0 else left_rz[:, :1])
        right_fit = _arc_fit(right_rz[:, right_near] if right_near.sum() > 0 else right_rz[:, :1])
    else:
        left_fit = _arc_fit(left_rz) if left_rz.shape[1] > 0 else None
        right_fit = _arc_fit(right_rz) if right_rz.shape[1] > 0 else None

    if left_fit is not None:
        result['left_angle'] = left_fit['angle']
        result['left_curvature'] = left_fit['curvature']
        if left_fit.get('contact') is not None:
            result['left_contact'] = Vector2(Q.T @ [left_fit['contact'], 0] + baseline.pt0)
        else:
            result['left_contact'] = Vector2(Q.T @ left_contact_rz + baseline.pt0) if left_contact_rz is not None else None
        if left_fit.get('arc_center') is not None:
            result['left_arc_center'] = Vector2(Q.T @ left_fit['arc_center'] + baseline.pt0)

    if right_fit is not None:
        if right_fit['angle'] is not None:
            result['right_angle'] = PI - right_fit['angle']
        result['right_curvature'] = right_fit['curvature']
        if right_fit.get('contact') is not None:
            result['right_contact'] = Vector2(Q.T @ [right_fit['contact'], 0] + baseline.pt0)
        else:
            result['right_contact'] = Vector2(Q.T @ right_contact_rz + baseline.pt0) if right_contact_rz is not None else None
        if right_fit.get('arc_center') is not None:
            result['right_arc_center'] = Vector2(Q.T @ right_fit['arc_center'] + baseline.pt0)

    return result


def _arc_fit(data):
    """
    Fit an arc (line or circle) to data points in (r, z) baseline coords.
    Returns angle from baseline, curvature, contact point, arc center.
    """
    if data.shape[1] < 2:
        return None

    line_res = line_fit(data)
    if line_res is None:
        return None

    # If line fit residuals are small, use line (curvature = 0)
    if (line_res['residuals'] < 1.0).all():
        curvature = 0.0
        arc_center = None
        angle = line_res['angle'] % PI
        pt = rotation_mat2d(angle) @ [0, line_res['rho']]
        unit = Vector2(np.cos(angle), np.sin(angle))
        if not np.isclose(unit.y, 0):
            grad = unit.x / unit.y
            contact = pt[0] - grad * pt[1]
        else:
            contact = None
        return {'contact': contact, 'angle': angle, 'curvature': curvature,
                'arc_center': arc_center, 'arclengths': None, 'residuals': line_res['residuals']}

    # Otherwise fit a circle
    circle_res = circle_fit(data)
    if circle_res is None:
        # Fall back to line
        angle = line_res['angle'] % PI
        return {'contact': None, 'angle': angle, 'curvature': 0.0, 'arc_center': None,
                'arclengths': None, 'residuals': line_res['residuals']}

    center = circle_res['center']
    radius = circle_res['radius']

    if center.y > radius:
        # Circle doesn't intersect baseline → fallback to line
        angle = line_res['angle'] % PI
        return {'contact': None, 'angle': angle, 'curvature': 0.0, 'arc_center': None,
                'arclengths': None, 'residuals': line_res['residuals']}

    curvature = 1.0 / radius
    l = np.sqrt(radius ** 2 - center.y ** 2)
    intersect1 = center.x - l
    intersect2 = center.x + l

    # Contact point = intersection closest to data
    if np.linalg.norm(data - [[intersect1], [0]], axis=0).min() < \
            np.linalg.norm(data - [[intersect2], [0]], axis=0).min():
        contact = intersect1
    else:
        contact = intersect2

    q = np.arctan2(center.y, center.x - contact)
    angle = (q + PI / 2) % PI

    # Sign convention
    if center.x > contact:
        curvature = -curvature
    elif center.x == contact:
        if data[0].mean() > contact:
            curvature = -curvature

    return {'contact': contact, 'angle': angle, 'curvature': curvature,
            'arc_center': center, 'arclengths': None, 'residuals': circle_res['residuals']}


# ===========================================================================
# 图像增强
# ===========================================================================
def enhance_image(img):
    """与 batch_process.py 相同的增强管线."""
    denoised = cv2.bilateralFilter(img, d=7, sigmaColor=75, sigmaSpace=75)
    lab = cv2.cvtColor(denoised, cv2.COLOR_BGR2LAB)
    l_ch, a_ch, b_ch = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
    l_eq = clahe.apply(l_ch)
    enhanced = cv2.cvtColor(cv2.merge((l_eq, a_ch, b_ch)), cv2.COLOR_LAB2BGR)
    blur = cv2.GaussianBlur(enhanced, (0, 0), sigmaX=2.0)
    enhanced = cv2.addWeighted(enhanced, 1.3, blur, -0.3, 0)
    return np.clip(enhanced, 0, 255).astype(np.uint8)


# ===========================================================================
# 自动基线检测
# ===========================================================================
def detect_baseline(gray, h, w):
    """Auto-detect the baseline (surface line) from gradient analysis."""
    sobel_y = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=5)
    sobel_abs = np.abs(sobel_y)
    row_means = np.mean(gray.astype(float), axis=1)
    row_means_s = uniform_filter1d(row_means, size=21)
    row_grad = np.diff(row_means_s)

    base_y_approx = (h // 4) + np.argmax(row_grad[h // 4:h - 30])
    y0, y1 = max(0, base_y_approx - 25), min(h, base_y_approx + 25)

    edge_pts = []
    for x in range(0, w, 3):
        col = sobel_abs[y0:y1, x]
        if np.max(col) > 20:
            edge_pts.append([x, y0 + np.argmax(col)])
    if len(edge_pts) < 10:
        for x in range(0, w, 5):
            edge_pts.append([x, base_y_approx])

    pts = np.array(edge_pts, dtype=np.float32)
    yv = pts[:, 1]
    ym = np.median(yv)
    y_mad = np.median(np.abs(yv - ym)) + 1e-8
    keep = np.abs(yv - ym) < max(3 * y_mad, 20)
    pts_in = pts[keep] if np.sum(keep) >= 5 else pts

    [vx, vy, x0, y0] = cv2.fitLine(pts_in, cv2.DIST_HUBER, 0, 0.01, 0.99)
    k = 0.0 if abs(vx) < 1e-8 else float((vy / vx)[0])
    b_val = float((y0 - k * x0)[0])
    if abs(k) > 0.2:
        k, b_val = 0.0, np.median(pts_in[:, 1])
    return k, b_val


# ===========================================================================
# 单张处理
# ===========================================================================
def process_single_image(image_path, idx, total):
    basename = os.path.basename(image_path)
    print(f"[{idx}/{total}] {basename} ...", end=" ", flush=True)

    result = {
        'filename': basename,
        'left_angle': None, 'right_angle': None, 'avg_angle': None,
        'left_curvature': None, 'right_curvature': None,
        'error': None,
    }

    try:
        # Use imdecode to handle Chinese paths on Windows
        with open(image_path, 'rb') as f:
            buf = np.frombuffer(f.read(), dtype=np.uint8)
        img = cv2.imdecode(buf, cv2.IMREAD_COLOR)
        if img is None:
            result['error'] = 'read'
            print("[FAIL] cannot read")
            return result

        h, w = img.shape[:2]
        enhanced = enhance_image(img)
        gray = cv2.cvtColor(enhanced, cv2.COLOR_BGR2GRAY)

        # 1. 基线检测
        k, b_val = detect_baseline(gray, h, w)
        baseline = Line2(Vector2(0.0, k * 0 + b_val), Vector2(float(w - 1), k * (w - 1) + b_val))

        # 2. OpenDrop 特征提取
        drop_points = extract_contact_angle_features(
            enhanced, baseline, inverted=False, thresh=0.3,
        )

        if drop_points.shape[1] < 10:
            result['error'] = 'no edge'
            print(f"[FAIL] insufficient edge points ({drop_points.shape[1]})")
            return result

        # 3. OpenDrop 接触角拟合
        fit = contact_angle_fit_opendrop(drop_points, baseline)

        left_deg = round(math.degrees(fit['left_angle']), 2) if fit['left_angle'] is not None else None
        right_deg = round(math.degrees(fit['right_angle']), 2) if fit['right_angle'] is not None else None

        if left_deg is None and right_deg is None:
            result['error'] = 'no angle'
            print("[FAIL] no angle computed")
            return result

        result['left_angle'] = left_deg
        result['right_angle'] = right_deg
        result['left_curvature'] = round(fit['left_curvature'], 6) if fit['left_curvature'] is not None else None
        result['right_curvature'] = round(fit['right_curvature'], 6) if fit['right_curvature'] is not None else None

        angles = [a for a in (left_deg, right_deg) if a is not None]
        result['avg_angle'] = round(sum(angles) / len(angles), 2)

        parts = []
        if left_deg is not None: parts.append(f"L={left_deg:.1f}°")
        if right_deg is not None: parts.append(f"R={right_deg:.1f}°")
        print(f"[OK] avg={result['avg_angle']:.1f}° ({' / '.join(parts)})")

    except Exception as e:
        result['error'] = str(e)[:120]
        print(f"[FAIL] {e}")

    return result


# ===========================================================================
# 写入 Excel
# ===========================================================================
def write_xlsx(results, elapsed, total):
    wb = openpyxl.Workbook()

    hf = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    hfl = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center')
    ca = Alignment(horizontal='center', vertical='center')
    tb = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'),
    )
    ef = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    # Sheet 1: Results
    ws1 = wb.active
    ws1.title = 'Results'
    headers = ['序号', '文件名', '左侧接触角(°)', '右侧接触角(°)', '平均接触角(°)',
               '左侧曲率', '右侧曲率', '错误']
    widths = [6, 32, 16, 16, 16, 14, 14, 22]

    for ci, (hdr, wid) in enumerate(zip(headers, widths), 1):
        c = ws1.cell(row=1, column=ci, value=hdr)
        c.font, c.fill, c.alignment, c.border = hf, hfl, ha, tb
        ws1.column_dimensions[get_column_letter(ci)].width = wid

    for ri, r in enumerate(results, 2):
        err = bool(r['error'])
        vals = [ri - 1, r['filename'],
                r['left_angle'] if not err else '', r['right_angle'] if not err else '',
                r['avg_angle'] if not err else '',
                r['left_curvature'] if not err else '', r['right_curvature'] if not err else '',
                r.get('error', '')]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            c.border, c.alignment = tb, ca
            if err: c.fill = ef

    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:H{len(results) + 1}"

    # Sheet 2: Statistics
    ws2 = wb.create_sheet('Statistics')
    valid = [r for r in results if not r['error']]
    angles = [r['avg_angle'] for r in valid if r['avg_angle'] is not None]
    left_a = [r['left_angle'] for r in valid if r['left_angle'] is not None]
    right_a = [r['right_angle'] for r in valid if r['right_angle'] is not None]

    stats = [
        ['指标', '数值'],
        ['算法', 'OpenDrop (contact_angle_fit)'],
        ['总数', total], ['成功', len(valid)], ['失败', total - len(valid)],
        ['成功率', f'{len(valid)/total*100:.1f}%' if total else '0%'],
        ['耗时', f'{elapsed:.1f}s'],
        ['', ''],
        ['--- 平均接触角 ---', ''],
        ['左侧平均值', f'{np.mean(left_a):.2f}°' if left_a else 'N/A'],
        ['右侧平均值', f'{np.mean(right_a):.2f}°' if right_a else 'N/A'],
        ['总体平均值', f'{np.mean(angles):.2f}°' if angles else 'N/A'],
        ['总体标准差', f'{np.std(angles):.2f}°' if angles else 'N/A'],
        ['最小值', f'{np.min(angles):.2f}°' if angles else 'N/A'],
        ['最大值', f'{np.max(angles):.2f}°' if angles else 'N/A'],
        ['中位数', f'{np.median(angles):.2f}°' if angles else 'N/A'],
    ]
    for ri, row in enumerate(stats, 1):
        for ci, v in enumerate(row, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.border, c.alignment = tb, ca
            if ri == 1: c.font, c.fill = hf, hfl
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 22

    wb.save(OUTPUT_XLSX)
    print(f"\n[OK] 结果已保存至: {OUTPUT_XLSX}")


# ===========================================================================
# Main
# ===========================================================================
def main():
    if not os.path.isdir(IMAGE_DIR):
        print(f"[ERROR] 图片目录不存在: {IMAGE_DIR}")
        sys.exit(1)

    files = sorted([f for f in os.listdir(IMAGE_DIR)
                    if f.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tif', '.tiff'))])
    if not files:
        print("[ERROR] 未找到图片"); sys.exit(1)

    total = len(files)
    print(f"{'=' * 60}")
    print(f"  OpenDrop 接触角批量测量")
    print(f"  图片数量: {total}")
    print(f"  算法: extract_contact_angle_features + contact_angle_fit")
    print(f"  (基于 OpenDrop - opendrop.dev)")
    print(f"{'=' * 60}\n")

    results = []
    t0 = time.time()
    for idx, fn in enumerate(files, 1):
        results.append(process_single_image(os.path.join(IMAGE_DIR, fn), idx, total))

    elapsed = time.time() - t0
    n_ok = sum(1 for r in results if not r['error'])

    print(f"\n{'=' * 60}")
    print(f"  完成! 耗时 {elapsed:.1f}s | 成功 {n_ok}/{total}")
    print(f"{'=' * 60}")

    write_xlsx(results, elapsed, total)

    valid = [r for r in results if not r['error']]
    if valid:
        angles = [r['avg_angle'] for r in valid if r['avg_angle'] is not None]
        if angles:
            print(f"\n  接触角: 均值={np.mean(angles):.2f}° ± {np.std(angles):.2f}°")
            print(f"  范围: [{np.min(angles):.1f}° ~ {np.max(angles):.1f}°]")
            print(f"  中位数: {np.median(angles):.2f}°")


if __name__ == '__main__':
    main()
